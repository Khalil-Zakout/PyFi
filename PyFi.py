from MoneyManagement import MoneyManager
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import datetime
import calendar

current_month = datetime.datetime.now().month
current_year = datetime.datetime.now().year
current_month_name = calendar.month_name[current_month]
sheet_name = f"{current_month_name}-{current_year}"

book = load_workbook("Expenses.xlsx")

if sheet_name not in book.sheetnames:
    sheet = book.create_sheet(sheet_name)
    sheet = book[sheet_name]

    # Setting Up Rows & Columns Names:
    sheet.cell(row=1, column=1).value = "Category"
    sheet.cell(row=1, column=1).font = Font(bold=True)

    sheet.cell(row=1, column=2).value = "Item"
    sheet.cell(row=1, column=2).font = Font(bold=True)

    sheet.cell(row=1, column=3).value = "Price"
    sheet.cell(row=1, column=3).font = Font(bold=True)

    sheet.cell(row=1, column=4).value = "Remaining Income"
    sheet.cell(row=1, column=4).font = Font(bold=True)

    sheet.cell(row=1, column=5).value = "Full Income"
    sheet.cell(row=1, column=5).font = Font(bold=True)

    sheet.cell(row=1, column=6).value = "Saving Goal"
    sheet.cell(row=1, column=6).font = Font(bold=True)

    book.save("Expenses.xlsx")
else:
    sheet = book.active

print("WELCOME To PyFi <3\n")

income_cell = sheet.cell(row=2, column=5)
saving_goal_cell = sheet.cell(row=2, column=6)

if income_cell.value is None:
    income = int(input("Enter Your Month's Income: "))
    income_cell.value = income
    book.save("Expenses.xlsx")

if saving_goal_cell.value is None:
    saving_goal = int(input("Enter Your Saving Goal: "))
    saving_goal_cell.value = saving_goal
    book.save("Expenses.xlsx")

user1 = MoneyManager(income_cell, sheet)

print("\n1) Add Expenses \n2) Reports \n3) Calculate Saving Goal \n")

while True:
    functionality = input("Choose What You Want To Do (Press Q to Exit): ")

    if functionality.lower() == 'q':
        print("\nBye")
        break

    if functionality == '1':
        category = input("Enter The Item Category: ")
        item = input("Enter The Name of The Item: ")
        cost = float(input("Enter The Cost of The Item: "))

        expense = {"Category": category, "Item": item, "Amount": float(cost)}

        user1.add_expenses(expense, user1.sheet)

    elif functionality == '2':
        user1.reports(user1.sheet)

    elif functionality == '3':
        item_price = float(input("Enter The Price of The Item: "))
        duration = int(input("Enter The Number of Months You Want The Item By: "))
        user1.calculate_savings_goal(item_price, duration)

    else:
        print("Please Try a Valid Option !!\n")

    print("\n1) Add Expenses \n2) Reports \n3) Calculate Saving Goal \n")
